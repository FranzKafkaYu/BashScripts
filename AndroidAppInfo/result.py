'''
Author: FranzKafka
Date: 2024-11-11 09:21:03
LastEditTime: 2024-11-11 09:24:57
LastEditors: Franzkafka
Description: 
FilePath: '/AppInspector/result.py'
'''

import re
from datetime import datetime
from typing import Dict, List, Any
import matplotlib.pyplot as plt
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io

class PerformanceParser:
    def __init__(self):
        self.result: Dict[str, Any] = {
            'timestamps': [],
            'app_info': {
                'package_name': '',
                'location': '',
                'size': '',
            },
            'permissions': {
                'requested': [],
                'install': {},
                'runtime': {}
            },
            'performance_data': defaultdict(list),
            'memory_details': {
                'java_heap': [],
                'native_heap': [],
                'code': [],
                'stack': [],
                'graphics': []
            }
        }
        self.current_timestamp = None
        # 从 result.txt 中解析得到的目标进程名（target process name）
        self.process_name = None

    def parse_file(self, file_path: str) -> Dict[str, Any]:
        """解析性能数据文件"""
        in_permission_section = False
        in_install_permission = False
        in_runtime_permission = False
        last_timestamp = None
        
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.readlines()
                in_meminfo_section = False
                memory_values = {}
                # 标记是否处于整体 CPU 统计区间（get cpu usage begin/end）
                in_cpu_section = False
                # 标记是否处于 each thread usage 区间
                in_thread_usage_section = False
                
                for line in content:
                    line = line.strip()
                    
                    # ---------------- CPU 统计区间控制 ----------------
                    # 只统计 get cpu usage begin 与 get cpu usage end 之间，
                    # 且排除 each thread usage begin 与 each thread usage end 之间的数据
                    if 'get cpu usage begin' in line:
                        in_cpu_section = True
                        in_thread_usage_section = False
                        continue
                    elif 'get cpu usage end' in line:
                        in_cpu_section = False
                        in_thread_usage_section = False
                        continue
                    elif 'each thread usage begin' in line:
                        # 子线程区间不参与统计
                        in_thread_usage_section = True
                        continue
                    elif 'each thread usage end' in line:
                        in_thread_usage_section = False
                        continue
                    
                    # 解析包信息
                    if 'name' in line and 'location:' in line:
                        parts = line.split('location:')
                        if len(parts) == 2:
                            name_part = parts[0].replace('name', '').strip()
                            self.result['app_info']['package_name'] = name_part
                            self.result['app_info']['location'] = parts[1].strip()
                    elif 'name' in line and 'size:' in line:
                        size_match = re.search(r'size:(\d+[MKG]B?)', line)
                        if size_match:
                            self.result['app_info']['size'] = size_match.group(1)

                    # 解析目标进程名：target process name: xxx
                    if 'target process name:' in line:
                        parts = line.split('target process name:')
                        if len(parts) >= 2:
                            self.process_name = parts[1].strip()
                            # 如果还没有包名，可以同时记录到 app_info 中
                            if not self.result['app_info']['package_name']:
                                self.result['app_info']['package_name'] = self.process_name
                    
                    # 解析权限信息
                    if 'requested permissions:' in line:
                        in_permission_section = True
                        in_install_permission = False
                        in_runtime_permission = False
                        continue
                    elif 'install permissions:' in line:
                        in_permission_section = False
                        in_install_permission = True
                        in_runtime_permission = False
                        continue
                    elif 'runtime permissions:' in line:
                        in_permission_section = False
                        in_install_permission = False
                        in_runtime_permission = True
                        continue
                    
                    # 处理请求的权限
                    if in_permission_section and (line.startswith('android.') or line.startswith('ohos.')):
                        perm = line.strip()
                        if perm not in self.result['permissions']['requested']:
                            self.result['permissions']['requested'].append(perm)
                    
                    # 处理安装时权限
                    elif in_install_permission and ': granted=' in line:
                        perm_match = re.search(r'([\w.]+): granted=(true|false)', line)
                        if perm_match:
                            perm_name, status = perm_match.groups()
                            self.result['permissions']['install'][perm_name] = status == 'true'
                    
                    # 处理运行时权限
                    elif in_runtime_permission and ': granted=' in line:
                        perm_match = re.search(r'([\w.]+): granted=(true|false)', line)
                        if perm_match:
                            perm_name, status = perm_match.groups()
                            self.result['permissions']['runtime'][perm_name] = status == 'true'
                    
                    # 检测时间戳
                    if 'Time:' in line:
                        current_timestamp = line.split('Time:')[1].strip()
                        if current_timestamp != last_timestamp:
                            self.current_timestamp = current_timestamp
                            last_timestamp = current_timestamp
                            if self.current_timestamp not in self.result['timestamps']:
                                self.result['timestamps'].append(self.current_timestamp)
                    
                    # 检测内存信息部分
                    if 'App Summary' in line:
                        in_meminfo_section = True
                        continue

                    if in_meminfo_section:
                        # 解析内存详细信息
                        if 'Java Heap:' in line:
                            match = re.search(r'Java Heap:\s+(\d+)', line)
                            if match:
                                value = int(match.group(1))
                                self.result['memory_details']['java_heap'].append({
                                    'timestamp': self.current_timestamp,
                                    'value': value
                                })
                        elif 'Native Heap:' in line:
                            match = re.search(r'Native Heap:\s+(\d+)', line)
                            if match:
                                value = int(match.group(1))
                                self.result['memory_details']['native_heap'].append({
                                    'timestamp': self.current_timestamp,
                                    'value': value
                                })
                        elif 'Code:' in line:
                            match = re.search(r'Code:\s+(\d+)', line)
                            if match:
                                value = int(match.group(1))
                                self.result['memory_details']['code'].append({
                                    'timestamp': self.current_timestamp,
                                    'value': value
                                })
                        elif 'Stack:' in line:
                            match = re.search(r'Stack:\s+(\d+)', line)
                            if match:
                                value = int(match.group(1))
                                self.result['memory_details']['stack'].append({
                                    'timestamp': self.current_timestamp,
                                    'value': value
                                })
                        elif 'Graphics:' in line:
                            match = re.search(r'Graphics:\s+(\d+)', line)
                            if match:
                                value = int(match.group(1))
                                self.result['memory_details']['graphics'].append({
                                    'timestamp': self.current_timestamp,
                                    'value': value
                                })
                        elif 'TOTAL PSS:' in line:
                            match = re.search(r'TOTAL PSS:\s+(\d+)', line)
                            if match and self.current_timestamp:
                                total_pss = int(match.group(1))
                                self.result['performance_data']['memory'].append({
                                    'timestamp': self.current_timestamp,
                                    'value': total_pss
                                })

                    # 解析CPU信息
                    # 只在整体 CPU 区间内统计进程的 CPU 占用，排除子线程明细区间
                    if (
                        in_cpu_section
                        and not in_thread_usage_section
                        and self.current_timestamp
                        and self.process_name
                        and self.process_name in line
                        and 'grep' not in line
                    ):
                        try:
                            parts = line.split()
                            if len(parts) >= 9:
                                cpu_usage = float(parts[8].strip('[]%'))
                                existing_data = next(
                                    (item for item in self.result['performance_data']['cpu'] 
                                     if item['timestamp'] == self.current_timestamp), 
                                    None
                                )
                                if existing_data:
                                    existing_data['value'] = max(existing_data['value'], cpu_usage)
                                else:
                                    self.result['performance_data']['cpu'].append({
                                        'timestamp': self.current_timestamp,
                                        'value': cpu_usage
                                    })
                        except (ValueError, IndexError) as e:
                            print(f"Error parsing CPU usage: {e}")
                            
                    # 解析GPU信息
                    if self.current_timestamp and 'Total GPU memory usage' in line:
                        try:
                            parts = line.split(':')
                            if len(parts) >= 2:
                                gpu_usage = float(parts[1].strip())
                                existing_data = next(
                                    (item for item in self.result['performance_data']['gpu'] 
                                     if item['timestamp'] == self.current_timestamp), 
                                    None
                                )
                                if existing_data:
                                    existing_data['value'] = max(existing_data['value'], gpu_usage)
                                else:
                                    self.result['performance_data']['gpu'].append({
                                        'timestamp': self.current_timestamp,
                                        'value': gpu_usage
                                    })
                        except (ValueError, IndexError) as e:
                            print(f"Error parsing GPU usage: {e}")
                
                return self.result
                
        except Exception as e:
            print(f"Error parsing file: {e}")
            return self.result

def plot_performance_charts(result) -> tuple:
    """绘制CPU、GPU、内存图表，返回三个图表的二进制数据"""
    plt.style.use('bmh')
    
    # 创建CPU图表
    plt.style.use('bmh')
    
    def format_timestamp(timestamp):
        try:
            dt = datetime.strptime(timestamp, "%a %b %d %H:%M:%S %Z-%Y")
            return dt.strftime("%H:%M:%S")
        except:
            return timestamp
    
    # 创建CPU图表
    plt.figure(figsize=(10, 6))
    cpu_data = result['performance_data']['cpu']
    
    # 检查是否有CPU数据
    if not cpu_data:
        plt.text(0.5, 0.5, 'No CPU data available', 
                horizontalalignment='center', verticalalignment='center')
        plt.title('CPU Usage Over Time', fontsize=14, pad=15)
    else:
        cpu_timestamps = [x['timestamp'] for x in cpu_data]
        cpu_values = [x['value'] for x in cpu_data]
        formatted_cpu_timestamps = [format_timestamp(ts) for ts in cpu_timestamps]
        
        plt.plot(range(len(cpu_values)), cpu_values, 'b-o', linewidth=2, markersize=6)
        plt.title('CPU Usage Over Time', fontsize=14, pad=15)
        plt.xlabel('Time', fontsize=12)
        plt.ylabel('CPU Usage (%)', fontsize=12)
        plt.grid(True, linestyle='--', alpha=0.7)
        
        if len(cpu_timestamps) > 0:
            num_ticks = min(10, len(cpu_timestamps))
            tick_indices = list(range(0, len(cpu_timestamps), max(1, len(cpu_timestamps) // num_ticks)))
            if tick_indices[-1] != len(cpu_timestamps) - 1:
                tick_indices.append(len(cpu_timestamps) - 1)
            plt.xticks(tick_indices, [formatted_cpu_timestamps[i] for i in tick_indices], rotation=45, ha='right')
    
    # 保存CPU图表到内存
    cpu_img_data = io.BytesIO()
    plt.savefig(cpu_img_data, format='png', bbox_inches='tight', dpi=100)
    plt.close()
    
    # 保存GPU图表到内存
    plt.figure(figsize=(10, 6))
    gpu_data = result['performance_data']['gpu']
    gpu_timestamps = [x['timestamp'] for x in gpu_data]
    gpu_values = [x['value'] for x in gpu_data]
    formatted_gpu_timestamps = [format_timestamp(ts) for ts in gpu_timestamps]
    
    plt.plot(range(len(gpu_values)), gpu_values, 'b-o', linewidth=2, markersize=6)
    plt.title('GPU Usage Over Time', fontsize=14, pad=15)
    plt.xlabel('Time', fontsize=12)
    plt.ylabel('GPU Usage (MB)', fontsize=12)
    plt.grid(True, linestyle='--', alpha=0.7)
    
    # 设置x轴标签
    num_ticks = min(10, len(gpu_timestamps))
    tick_indices = list(range(0, len(gpu_timestamps), max(1, len(gpu_timestamps) // num_ticks)))
    if tick_indices[-1] != len(gpu_timestamps) - 1:
        tick_indices.append(len(gpu_timestamps) - 1)
    
    plt.xticks(tick_indices, [formatted_gpu_timestamps[i] for i in tick_indices], rotation=45, ha='right')
    
    # 保存GPU图表到内存
    gpu_img_data = io.BytesIO()
    plt.savefig(gpu_img_data, format='png', bbox_inches='tight', dpi=100)
    plt.close()

    # 创建内存图表
    plt.figure(figsize=(10, 6))
    memory_types = {
        'java_heap': ('Java Heap', 'b-'),
        'native_heap': ('Native Heap', 'g-'),
        'code': ('Code', 'r-'),
        'stack': ('Stack', 'c-'),
        'graphics': ('Graphics', 'm-')
    }
    
    for mem_type, (label, style) in memory_types.items():
        if result['memory_details'][mem_type]:
            values = [x['value']/1024 for x in result['memory_details'][mem_type]]
            timestamps = [x['timestamp'] for x in result['memory_details'][mem_type]]
            plt.plot(range(len(values)), values, style, linewidth=1.5, label=label, alpha=0.7)
    
    memory_data = result['performance_data']['memory']
    memory_values = [x['value']/1024 for x in memory_data]
    plt.plot(range(len(memory_values)), memory_values, 'k-', linewidth=2, label='Total PSS', alpha=0.5)
    
    plt.title('Memory Usage Over Time', fontsize=14, pad=15)
    plt.xlabel('Time', fontsize=12)
    plt.ylabel('Memory Usage (MB)', fontsize=12)
    plt.grid(True, linestyle='--', alpha=0.7)
    plt.legend(loc='upper right')
    
    # 保存内存图表到内存
    memory_img_data = io.BytesIO()
    plt.savefig(memory_img_data, format='png', bbox_inches='tight', dpi=100)
    plt.close()


    return cpu_img_data, gpu_img_data, memory_img_data

def create_excel_report(result, output_file='performance_report.xlsx'):
    """生成Excel报告"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Performance Report"
    
    # 设置样式
    title_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 应用基本信息
    ws['A1'] = "Application Information"
    ws['A1'].font = title_font
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:B1')
    
    headers = [
        ('Package Name', result['app_info']['package_name']),
        ('APK Size', result['app_info']['size']),
        ('Location', result['app_info']['location'])
    ]
    
    for i, (header, value) in enumerate(headers, start=2):
        ws[f'A{i}'] = header
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)
        ws[f'A{i}'].border = border
        ws[f'B{i}'].border = border
    
    # 权限信息
    current_row = 6
    ws[f'A{current_row}'] = "Permissions"
    ws[f'A{current_row}'].font = title_font
    ws[f'A{current_row}'].fill = header_fill
    ws.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 1
    
    # 请求的权限
    if result['permissions']['requested']:
        ws[f'A{current_row}'] = "Requested Permissions"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'A{current_row}'].fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        ws.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1
        
        for perm in result['permissions']['requested']:
            ws[f'A{current_row}'] = perm
            ws[f'A{current_row}'].border = border
            ws.merge_cells(f'A{current_row}:B{current_row}')
            current_row += 1
    
    # 安装时权限
    if result['permissions']['install']:
        current_row += 1
        ws[f'A{current_row}'] = "Install-time Permissions"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'A{current_row}'].fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        ws.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1
        
        for perm, status in result['permissions']['install'].items():
            ws[f'A{current_row}'] = perm
            ws[f'B{current_row}'] = str(status)
            ws[f'A{current_row}'].border = border
            ws[f'B{current_row}'].border = border
            current_row += 1
    
    # 运行时权限
    if result['permissions']['runtime']:
        current_row += 1
        ws[f'A{current_row}'] = "Runtime Permissions"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'A{current_row}'].fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        ws.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1
        
        for perm, status in result['permissions']['runtime'].items():
            ws[f'A{current_row}'] = perm
            ws[f'B{current_row}'] = str(status)
            ws[f'A{current_row}'].border = border
            ws[f'B{current_row}'].border = border
            current_row += 1
    
    # 性能数据摘要
    current_row += 2
    ws[f'A{current_row}'] = "Performance Summary"
    ws[f'A{current_row}'].font = title_font
    ws[f'A{current_row}'].fill = header_fill
    ws.merge_cells(f'A{current_row}:B{current_row}')
    current_row += 1
    
    if result['performance_data']['cpu']:
        cpu_data = result['performance_data']['cpu']
        avg_cpu = sum(x['value'] for x in cpu_data) / len(cpu_data)
        max_cpu = max(x['value'] for x in cpu_data)
        min_cpu = min(x['value'] for x in cpu_data)
        
        ws[f'A{current_row}'] = "CPU Usage"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'A{current_row}'].fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        ws.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1
        
        for label, value in [("Average", f"{avg_cpu:.2f}%"), 
                           ("Maximum", f"{max_cpu:.2f}%"), 
                           ("Minimum", f"{min_cpu:.2f}%")]:
            ws[f'A{current_row}'] = label
            ws[f'B{current_row}'] = value
            ws[f'A{current_row}'].border = border
            ws[f'B{current_row}'].border = border
            current_row += 1
    
    if result['performance_data']['memory']:
        current_row += 1
        memory_data = result['performance_data']['memory']
        avg_memory = sum(x['value'] for x in memory_data) / len(memory_data)
        max_memory = max(x['value'] for x in memory_data)
        min_memory = min(x['value'] for x in memory_data)
        
        ws[f'A{current_row}'] = "Memory Usage (Total PSS)"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'A{current_row}'].fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        ws.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1
        
        for label, value in [("Average", f"{avg_memory/1024:.2f} MB"), 
                           ("Maximum", f"{max_memory/1024:.2f} MB"), 
                           ("Minimum", f"{min_memory/1024:.2f} MB")]:
            ws[f'A{current_row}'] = label
            ws[f'B{current_row}'] = value
            ws[f'A{current_row}'].border = border
            ws[f'B{current_row}'].border = border
            current_row += 1
    
    # 添加性能图表
    cpu_img_data, gpu_img_data, memory_img_data = plot_performance_charts(result)
    
    # 重置图片数据指针
    cpu_img_data.seek(0)
    gpu_img_data.seek(0)
    memory_img_data.seek(0)
    
    # 创建并插入图片
    cpu_img = Image(cpu_img_data)
    gpu_img = Image(gpu_img_data)
    memory_img = Image(memory_img_data)
    
    # 调整图片大小
    scale_factor = 0.7
    cpu_img.width = int(cpu_img.width * scale_factor)
    cpu_img.height = int(cpu_img.height * scale_factor)
    gpu_img.width = int(gpu_img.width * scale_factor)
    gpu_img.height = int(gpu_img.height * scale_factor)
    memory_img.width = int(memory_img.width * scale_factor)
    memory_img.height = int(memory_img.height * scale_factor)
    
    # 添加图表标题
    ws['D1'] = "Performance Charts"
    ws['D1'].font = title_font
    ws['D1'].fill = header_fill
    
    # 插入图片
    ws.add_image(cpu_img, 'D2')
    ws.add_image(gpu_img, f'D{1 + int(cpu_img.height/15)}') 
    ws.add_image(memory_img, f'D{1+ int(cpu_img.height/15) * 2}')  # 15是每行的近似高度
    
    # 调整列宽
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 40
    
    # 保存Excel文件
    wb.save(output_file)

def main():
    parser = PerformanceParser()
    result = parser.parse_file('result.txt')
    
    # 生成Excel报告
    create_excel_report(result)
    print("\nPerformance report has been saved to 'performance_report.xlsx'")

if __name__ == "__main__":
    main()